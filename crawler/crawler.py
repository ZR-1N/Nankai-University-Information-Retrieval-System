"""
主爬虫模块 - 协调下载、解析、存储
优化版 v3.0.0：优先级队列 + 多源均衡爬取 + URL元数据传递 + page_role 记录
"""
import os
import json
import time
from datetime import datetime
from tqdm import tqdm

from config.settings import SEED_URLS_FILE
from crawler.url_manager import (
    URLManager, URLEntry, is_allowed_domain,
    is_attachment_url, classify_url,
)
from crawler.downloader import Downloader
from crawler.parser import HTMLParser
from crawler.snapshot import SnapshotManager
from crawler.utils import (
    generate_doc_id, save_metadata, compute_stats, save_stats,
    extract_site_name, clean_text,
)


class Crawler:
    """通用爬虫：优先级 BFS + 多源均衡爬取，支持断点续爬"""

    def __init__(self, limit: int = 100, resume: bool = False,
                 delay: float = 1.0, fresh: bool = False,
                 balanced: bool = True):
        self.limit = limit
        self.delay = delay
        self.resume = resume
        self.fresh = fresh
        self.balanced = balanced
        self.counter = 0
        self.failed_urls = 0
        self.duplicate_urls = 0
        self.start_time = None

        # 加载配置
        with open(SEED_URLS_FILE, "r", encoding="utf-8") as f:
            self.config = json.load(f)

        self.allowed_domains = self.config.get("allowed_domains", ["nankai.edu.cn"])
        self.sources = self.config.get("sources", [])

        # 初始化各组件
        max_source_ratio = 0.30 if balanced else 0.99
        self.url_manager = URLManager(
            balanced=balanced,
            max_source_ratio=max_source_ratio,
        )

        # 构建 seed URL→文章类别映射（依赖 url_manager）
        self._seed_map = self._build_seed_map()
        self.downloader = Downloader(delay=delay)
        self.snapshot_manager = SnapshotManager(max_snapshots=50)
        self.pbar = None

        # fresh 模式：清除旧数据
        if fresh:
            self._clear_old_data()

        # 断点续爬
        if resume:
            loaded = self.url_manager.load_checkpoint()
            if loaded:
                from crawler.utils import load_metadata
                existing = load_metadata()
                self.counter = len(existing)
                # 从断点恢复 source_crawled 计数
                print(f"[断点续爬] 已恢复 {self.counter} 条记录，"
                      f"队列剩余 {self.url_manager.queue_size} 个 URL")
            else:
                print("[断点续爬] 未找到断点文件，从头开始爬取")

    def _build_seed_map(self) -> dict:
        """构建 URL → {source_site, category, page_role} 映射"""
        seed_map = {}
        for source in self.sources:
            source_name = source.get("name", "")
            for item in source.get("seed_urls", []):
                if isinstance(item, str):
                    url = item
                    category = ""
                else:
                    url = item.get("url", "")
                    category = item.get("category", "")
                seed_map[self.url_manager._normalize_url(url)] = {
                    "source_site": source_name,
                    "category": category,
                }
        return seed_map

    def _clear_old_data(self):
        """清除旧数据文件"""
        from config.settings import METADATA_FILE, CRAWL_STATS_FILE
        import glob
        for path in [METADATA_FILE, CRAWL_STATS_FILE,
                     self.url_manager.checkpoint_path]:
            if os.path.exists(path):
                os.remove(path)
        # 清除旧 checkpoint
        for old in glob.glob("data/url_manager_checkpoint*"):
            os.remove(old)
        print("[fresh] 已清除旧数据")

    def run(self):
        """执行爬取"""
        self.start_time = time.time()
        mode_str = "均衡" if self.balanced else "优先级"
        print(f"开始爬取（{mode_str}模式），限制 {self.limit} 个页面/文档...")

        # 初始化种子 URL
        if self.counter == 0:
            self.url_manager.add_seed_urls(self.sources)

        with tqdm(total=self.limit, desc="爬取进度", unit="页") as self.pbar:
            self.pbar.update(self.counter)

            while self.counter < self.limit:
                entry = self.url_manager.get_next(total_target=self.limit)
                if entry is None:
                    print(f"\n队列已空！已爬取 {self.counter} 个文档。")
                    break

                url = entry.url

                # 域名检查
                if not is_allowed_domain(url, self.allowed_domains):
                    continue

                # 已访问检查
                if self.url_manager.is_visited(url):
                    continue

                self.url_manager.mark_visited(url)

                # 根据 URL 类型分发处理
                if is_attachment_url(url):
                    self._process_attachment(url, entry)
                else:
                    self._process_html(url, entry)

                # 定期保存断点
                if self.counter % 50 == 0:
                    self.url_manager.save_checkpoint()
                    self._update_stats()

        # 最终保存
        self.url_manager.save_checkpoint()
        self._update_stats()
        self._print_summary()

    def _process_html(self, url: str, entry: URLEntry):
        """处理 HTML 页面"""
        html, error = self.downloader.fetch_html(url)
        if error:
            self.failed_urls += 1
            return

        if not html:
            return

        # 获取 seed 配置信息
        normalized = self.url_manager._normalize_url(url)
        seed_info = self._seed_map.get(normalized, {})
        source_site = entry.source_site or seed_info.get("source_site", "")
        if not source_site:
            source_site = extract_site_name(url, self.sources)
        seed_category = entry.category or seed_info.get("category", "")

        # 解析页面
        parser = HTMLParser(
            html, url,
            anchor_text=entry.anchor_text,
            source_site=source_site,
            seed_category=seed_category,
        )

        title = parser.extract_title()
        content = parser.extract_content()
        content = clean_text(content)
        summary = parser.extract_summary(content)
        publish_time = parser.extract_publish_time()
        page_role = parser.classify_page_role()

        # 内容质量评估
        content_quality = parser.assess_content_quality(content, title, page_role)

        # 去重
        content_hash = URLManager.make_content_hash(title, content)
        if not self.url_manager.add_content_hash(content_hash):
            self.duplicate_urls += 1
            return

        # 提取链接并加入队列
        link_entries = parser.extract_links()
        for link_info in link_entries:
            link_url = link_info["url"]
            if not is_allowed_domain(link_url, self.allowed_domains):
                continue
            is_detail = link_info["is_detail"]
            self.url_manager.add_url(
                link_url,
                anchor_text=link_info["anchor_text"],
                source_site=source_site,
                category=seed_category,
                parent_url=url,
                is_detail=is_detail,
            )

        # 分页链接
        pagination_links = parser.extract_pagination_links()
        for link_url in pagination_links:
            if is_allowed_domain(link_url, self.allowed_domains):
                self.url_manager.add_url(
                    link_url,
                    source_site=source_site,
                    category=seed_category,
                    parent_url=url,
                    priority=2,  # list
                )

        # 附件链接（高优先级）
        attachment_links = parser.extract_attachment_links()
        for link_url in attachment_links:
            if is_allowed_domain(link_url, self.allowed_domains):
                self.url_manager.add_url(
                    link_url,
                    anchor_text=parser._get_attachment_anchor_text(link_url),
                    source_site=source_site,
                    category=seed_category,
                    parent_url=url,
                    priority=0,  # document
                )

        # 检查该 source 是否队列耗尽（用于动态配额重分配）
        if source_site and not self.url_manager._source_has_entries(source_site):
            if not self.url_manager.is_source_depleted(source_site):
                self.url_manager.mark_source_depleted(source_site)
                active_count = sum(1 for s in self.url_manager._source_round_order
                                  if not self.url_manager.is_source_depleted(s))
                if active_count == 0:
                    active_count = len(self.url_manager._source_round_order)
                # 静默记录，不打印（避免刷屏）；_print_summary 会汇总

        # 生成 doc_id
        self.counter += 1
        doc_id = generate_doc_id(self.counter, "html")

        # 保存原始 HTML
        local_path = self.downloader.save_html(html, doc_id)

        # 尝试保存快照
        snap_meta = {
            "url": url,
            "title": title,
            "content": content,
            "source_site": source_site,
            "file_type": "html",
            "publish_time": publish_time,
        }
        snapshot_path = ""
        if self.snapshot_manager.should_save_snapshot(snap_meta):
            snapshot_path = self.snapshot_manager.save_snapshot(
                html, doc_id, url
            )

        # 构建元数据
        metadata = {
            "doc_id": doc_id,
            "url": url,
            "title": title,
            "content": content,
            "content_length": len(content),
            "content_quality": content_quality,
            "summary": summary,
            "source_site": source_site,
            "category": seed_category or entry.category,
            "publish_time": publish_time,
            "file_type": "html",
            "page_role": page_role,
            "local_path": local_path,
            "download_path": "",
            "snapshot_path": snapshot_path,
            "content_hash": content_hash,
            "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "parse_status": "ok",
            "parse_error": "",
        }

        save_metadata(metadata)

        # 记录 source_site 成功爬取
        self.url_manager.record_crawled(source_site)
        self.pbar.update(1)

    def _process_attachment(self, url: str, entry: URLEntry):
        """处理附件下载"""
        local_path, content_type, error = self.downloader.download_file(url)
        if error:
            self.failed_urls += 1
            return

        if not local_path:
            return

        self.counter += 1
        file_type = classify_url(url)
        doc_id = generate_doc_id(self.counter, file_type)

        filename = os.path.basename(local_path)
        source_site = entry.source_site or extract_site_name(url, self.sources)

        # 解析文档内容（返回 (content, status, error_msg)）
        content = ""
        parse_status = "ok"
        parse_error = ""
        if file_type == "pdf":
            content, parse_status, parse_error = self._parse_pdf(local_path)
        elif file_type in ("docx", "doc"):
            content, parse_status, parse_error = self._parse_docx(local_path)
        elif file_type in ("xlsx", "xls"):
            content, parse_status, parse_error = self._parse_xlsx(local_path)
        # zip/rar 不解析文本，标记为 skipped
        elif file_type in ("zip", "rar"):
            parse_status = "skipped"
            parse_error = "ZIP/RAR 文件不解码文本内容"

        # 标题优先使用 anchor_text，其次文件名
        title = entry.anchor_text if entry.anchor_text else filename
        # 清理标题
        title = title.replace(".pdf", "").replace(".docx", "").replace(".doc", "")\
                     .replace(".xlsx", "").replace(".xls", "").strip()

        # 如果无法解析内容，用文件名+父页面信息作为 content
        if not content or not content.strip():
            content_parts = [title]
            if entry.anchor_text:
                content_parts.append(f"链接文本: {entry.anchor_text}")
            if entry.category:
                content_parts.append(f"分类: {entry.category}")
            content = " | ".join(content_parts)

        content_quality = "good" if len(content) > 50 else "short"
        if parse_status == "failed":
            # 解析失败但仍有元数据兜底
            content_quality = "short"

        metadata = {
            "doc_id": doc_id,
            "url": url,
            "title": title,
            "content": clean_text(content),
            "content_length": len(content),
            "content_quality": content_quality,
            "summary": clean_text(content[:200]) if content else filename,
            "source_site": source_site,
            "category": entry.category,
            "publish_time": "",
            "file_type": file_type,
            "page_role": "document",
            "local_path": local_path,
            "download_path": local_path,
            "parent_url": entry.parent_url,
            "parent_title": "",
            "snapshot_path": "",
            "content_hash": URLManager.make_content_hash(filename, content),
            "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "parse_status": parse_status,
            "parse_error": parse_error,
        }

        save_metadata(metadata)

        # 记录 source_site 成功爬取
        self.url_manager.record_crawled(source_site if source_site else "其他")
        self.pbar.update(1)

    def _parse_pdf(self, filepath: str) -> tuple[str, str, str]:
        """解析 PDF，返回 (content, status, error_msg)

        使用 os.dup2() 在文件描述符层面重定向 stderr (fd=2)，
        彻底捕获 MuPDF C 库写入 fd=2 的错误信息。
        """
        content = ""
        status = "ok"
        error_msg = ""

        # 确定错误日志路径
        log_dir = os.path.join(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))), "logs")
        os.makedirs(log_dir, exist_ok=True)
        error_log_path = os.path.join(log_dir, "document_parse_errors.log")

        # 使用 os.dup2() 在文件描述符层面重定向 stderr
        # 这能捕获 C 库（如 MuPDF）直接写入 fd=2 的输出
        stderr_fd = 2
        old_stderr_fd = os.dup(stderr_fd)  # 备份原 fd=2
        log_fd = os.open(error_log_path, os.O_WRONLY | os.O_CREAT | os.O_APPEND, 0o644)
        os.dup2(log_fd, stderr_fd)  # fd=2 → 日志文件
        os.close(log_fd)

        try:
            import fitz
            doc = fitz.open(filepath)
            texts = [page.get_text() for page in doc]
            doc.close()
            content = "\n".join(texts)
            if not content.strip():
                status = "failed"
                error_msg = "PyMuPDF extracted empty text"
        except Exception as e:
            status = "failed"
            error_msg = f"PyMuPDF error: {str(e)[:200]}"
            # fallback to pdfplumber
            try:
                import pdfplumber
                texts = []
                with pdfplumber.open(filepath) as pdf:
                    for page in pdf.pages:
                        t = page.extract_text()
                        if t:
                            texts.append(t)
                content = "\n".join(texts)
                if content.strip():
                    status = "ok"
                    error_msg = ""
            except Exception as e2:
                error_msg = f"pdfplumber also failed: {str(e2)[:200]}"
        finally:
            # 恢复原 stderr (fd=2)
            os.dup2(old_stderr_fd, stderr_fd)
            os.close(old_stderr_fd)

        # 记录解析错误（如果有的话）
        # 注意：MuPDF 的 C 层错误已经直接写入日志文件，这里记录 Python 层异常
        if error_msg:
            self._log_parse_error(filepath, "pdf", error_msg)

        return content, status, error_msg

    def _parse_docx(self, filepath: str) -> tuple[str, str, str]:
        """解析 DOCX，返回 (content, status, error_msg)"""
        content = ""
        status = "ok"
        error_msg = ""

        try:
            from docx import Document
            doc = Document(filepath)
            content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            if not content.strip():
                status = "failed"
                error_msg = "python-docx extracted empty text"
        except Exception as e:
            status = "failed"
            error_msg = f"python-docx error: {str(e)[:200]}"

        if error_msg:
            self._log_parse_error(filepath, "docx", error_msg)

        return content, status, error_msg

    def _parse_xlsx(self, filepath: str) -> tuple[str, str, str]:
        """解析 XLSX，返回 (content, status, error_msg)"""
        import warnings
        content = ""
        status = "ok"
        error_msg = ""

        try:
            import openpyxl
            # 抑制 openpyxl 的 Data Validation warning
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                warnings.simplefilter("ignore", DeprecationWarning)
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join(str(c) for c in row if c is not None)
                    if row_text.strip():
                        texts.append(row_text)
            wb.close()
            content = "\n".join(texts[:200])
            if not content.strip():
                status = "failed"
                error_msg = "openpyxl extracted empty text"
        except Exception as e:
            status = "failed"
            error_msg = f"openpyxl error: {str(e)[:200]}"

        if error_msg:
            self._log_parse_error(filepath, "xlsx", error_msg)

        return content, status, error_msg

    def _log_parse_error(self, filepath: str, file_type: str, error: str):
        """记录文档解析错误到日志文件"""
        import os
        log_dir = os.path.join(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))), "logs")
        os.makedirs(log_dir, exist_ok=True)
        log_path = os.path.join(log_dir, "document_parse_errors.log")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] [{file_type.upper()}] {filepath}\n")
            f.write(f"  Error: {error}\n")

    def _update_stats(self):
        # 注意：compute_stats(False) 不从旧文件恢复，避免爬取过程统计被覆盖
        stats = compute_stats(preserve_process_stats=False)
        stats["failed_urls"] = self.failed_urls
        stats["duplicate_urls"] = self.duplicate_urls
        if self.start_time:
            stats["crawl_elapsed"] = round(time.time() - self.start_time, 1)
            stats["crawl_speed"] = round(self.counter / max(stats["crawl_elapsed"], 0.1), 2)
        # 添加均衡爬取统计
        stats["balanced_mode"] = self.balanced
        stats["max_source_ratio"] = self.url_manager.max_source_ratio
        stats["source_crawled"] = self.url_manager.source_distribution()
        stats["source_quota_info"] = self.url_manager.get_source_quota_info(
            total_target=self.limit
        )
        save_stats(stats)

    def _print_summary(self):
        elapsed = time.time() - self.start_time
        queue_sizes = self.url_manager.queue_sizes_by_priority()
        source_dist = self.url_manager.source_distribution()

        print(f"\n{'=' * 60}")
        print(f"爬取完成！")
        print(f"  成功: {self.counter} 个文档")
        print(f"  失败: {self.failed_urls} 个 URL")
        print(f"  重复: {self.duplicate_urls} 个页面")
        print(f"  耗时: {elapsed:.1f}s")
        if elapsed > 0:
            print(f"  速度: {self.counter / elapsed:.1f} 页/秒")
        print(f"  队列分布: doc={queue_sizes.get(0,0)} detail={queue_sizes.get(1,0)} "
              f"list={queue_sizes.get(2,0)} portal={queue_sizes.get(3,0)}")

        # 打印各来源分布
        if source_dist:
            print(f"\n  Source 分布:")
            for site, count in sorted(source_dist.items(),
                                       key=lambda x: x[1], reverse=True):
                pct = round(count / max(self.counter, 1) * 100, 1)
                print(f"    {site}: {count} ({pct}%)")

        print(f"{'=' * 60}")
